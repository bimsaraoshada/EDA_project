from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


SRC_POP = "C:/Users/DELL/Downloads/Population_Table_western.xlsx"
SRC_EV = "C:/Users/DELL/Downloads/Western_Province_Data.xlsx"
SRC_CHG = "C:/Users/DELL/Downloads/Western_Province_EV_Chargers.xlsx"
OUT = "C:/Users/DELL/Downloads/Merged_dataset.xlsx"


def normalize_text(value):
    if value is None:
        return None
    text = str(value).replace("\n", " ").replace("\r", " ").strip()
    text = " ".join(text.split())
    if not text or text.lower() == "nan":
        return None
    return text


def normalize_key(value):
    text = normalize_text(value)
    if text is None:
        return None
    return text.casefold()


def normalize_district_key(value):
    text = normalize_text(value)
    if text is None:
        return None
    lowered = text.casefold()
    if lowered.endswith(" district"):
        lowered = lowered[: -len(" district")].strip()
    return lowered


def normalize_division_key(value):
    key = normalize_key(value)
    if key == "colomno":
        return "colombo"
    return key


def resolve_index(columns, candidates, label):
    lookup = {}
    for index, column in enumerate(columns):
        normalized = normalize_text(column)
        if normalized:
            lookup[normalized.casefold()] = index

    for candidate in candidates:
        resolved = lookup.get(candidate.casefold())
        if resolved is not None:
            return resolved

    available = [normalize_text(column) for column in columns]
    raise ValueError(f"Could not find a {label} column. Available columns: {available}")


def load_population_rows(path):
    workbook = load_workbook(path, data_only=True, read_only=True)
    sheet = workbook.active

    rows = []
    current_district = None

    for district, ds_division, population, *_ in sheet.iter_rows(min_row=8, values_only=True):
        district_text = normalize_text(district)
        ds_text = normalize_text(ds_division)
        population_value = population

        if district_text is None and ds_text is None:
            continue

        if district_text is not None:
            current_district = district_text

        if ds_text is None:
            continue

        # Skip national total rows like 'Sri Lanka'
        if (
            (district_text and district_text.casefold() == "sri lanka")
            or (ds_text and ds_text.casefold() == "sri lanka")
            or (current_district and current_district.casefold() == "sri lanka")
        ):
            continue

        rows.append(
            {
                "District": current_district,
                "Divisional Secretariat Division": ds_text,
                "Population": population_value,
            }
        )

    return rows


def load_ev_counts(path):
    workbook = load_workbook(path, data_only=True, read_only=True)
    sheet = workbook.active
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))

    district_index = resolve_index(header_row, ["District"], "EV district")

    count_index = None
    for candidate in ["Count", "EV Count", "Number of EVs"]:
        try:
            count_index = resolve_index(header_row, [candidate], "EV count")
            break
        except ValueError:
            continue

    ev_counts = defaultdict(int)
    for row in sheet.iter_rows(min_row=2, values_only=True):
        district_key = normalize_district_key(row[district_index] if district_index < len(row) else None)
        if district_key is None:
            continue

        if count_index is not None and count_index < len(row):
            raw_count = row[count_index]
            try:
                count_value = int(float(raw_count)) if raw_count is not None else 1
            except (TypeError, ValueError):
                count_value = 1
        else:
            count_value = 1

        ev_counts[district_key] += count_value

    return ev_counts


def load_charger_counts(path):
    workbook = load_workbook(path, data_only=True, read_only=True)
    sheet = workbook.active

    header_row = next(sheet.iter_rows(min_row=2, max_row=2, values_only=True))
    district_index = resolve_index(header_row, ["District"], "charger district")
    division_index = resolve_index(
        header_row,
        ["Divisional Secretary's Division", "Divisional Secratariat Division", "DS Division"],
        "charger DS division",
    )

    charger_counts = Counter()
    for row in sheet.iter_rows(min_row=3, values_only=True):
        if row is None:
            continue

        district_key = normalize_district_key(row[district_index] if district_index < len(row) else None)
        division_key = normalize_division_key(row[division_index] if division_index < len(row) else None)
        if district_key is None or division_key is None:
            continue

        charger_counts[(district_key, division_key)] += 1

    return charger_counts


def main():
    population_rows = load_population_rows(SRC_POP)
    ev_counts = load_ev_counts(SRC_EV)
    charger_counts = load_charger_counts(SRC_CHG)

    district_population = defaultdict(int)
    district_max_density = defaultdict(float)
    district_min_density = defaultdict(lambda: None)

    for row in population_rows:
        district = row["District"]
        population = row["Population"]
        if district is None or population in (None, ""):
            continue
        try:
            district_population[district] += float(population)
        except (TypeError, ValueError):
            continue

    for row in population_rows:
        district = row["District"]
        population = row["Population"]
        if district is None or population in (None, ""):
            continue
        try:
            density = float(population)
        except (TypeError, ValueError):
            continue
        current_min = district_min_density[district]
        current_max = district_max_density[district]
        if current_min is None or density < current_min:
            district_min_density[district] = density
        if density > current_max:
            district_max_density[district] = density

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Merged Data"

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    title_font = Font(name="Arial", bold=True, size=14)
    note_font = Font(name="Arial", italic=True, size=9, color="666666")
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    sheet["A1"] = "Merged EV Demand Table"
    sheet["A1"].font = title_font

    sheet["A3"].font = note_font

    headers = [
        "District",
        "Divisional Secretariat Division",
        "Population",
        "Area",
        "Population Share",
        "Urbanization Proxy",
        "Demand Score",
        "District EV Count",
        "Divisional Secretariat EV Demand",
        "Existing Charging Points in Divisional Secretariat",
        "Expected Charging Stations",
    ]

    header_row = 5
    for index, label in enumerate(headers, start=1):
        cell = sheet.cell(header_row, index, label)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    sheet.row_dimensions[header_row].height = 42

    widths = [14, 28, 12, 12, 15, 16, 18, 15, 18, 28, 20]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    data_start = header_row + 1
    data_end = data_start + len(population_rows) - 1

    for row_index, row in enumerate(population_rows, start=data_start):
        district = row["District"]
        division = row["Divisional Secretariat Division"]
        population = row["Population"]
        district_key = normalize_district_key(district)
        division_key = normalize_division_key(division)
        district_ev_total = ev_counts.get(district_key, 0)

        sheet.cell(row_index, 1, district)
        sheet.cell(row_index, 2, division)
        sheet.cell(row_index, 3, population)

        sheet.cell(row_index, 4, None)
        sheet.cell(row_index, 5, None)
        sheet.cell(row_index, 6, None)
        sheet.cell(row_index, 7, None)
        sheet.cell(row_index, 8, district_ev_total)
        sheet.cell(row_index, 9, None)
        sheet.cell(row_index, 10, charger_counts.get((district_key, division_key), 0))
        sheet.cell(row_index, 11, None)

    for row_index in range(data_start, data_end + 1):
        for column_index in range(1, 12):
            cell = sheet.cell(row_index, column_index)
            cell.font = Font(name="Arial")
            cell.border = border
            if column_index in (3, 4, 8, 9, 10, 11):
                cell.number_format = '#,##0'

    sheet.freeze_panes = f"A{data_start}"

    note_row = data_end + 2
    notes = [
        "Notes:",
        "- Population is imported from Population_Table_western.xlsx.",
        "- District EV Count is summed from Western_Province_Data.xlsx.",
        "- Existing charging points are counted from Western_Province_EV_Chargers.xlsx.",
        "- District EV Count and Divisional Secretariat EV Demand are repeated for every DS row in the same district.",
    ]
    for offset, text in enumerate(notes):
        cell = sheet.cell(note_row + offset, 1, text)
        cell.font = note_font if offset else Font(name="Arial", bold=True, size=9)

    try:
        workbook.save(OUT)
        saved_to = OUT
    except PermissionError:
        output_path = Path(OUT)
        for index in range(1, 1000):
            candidate = output_path.with_name(f"{output_path.stem}_{index}{output_path.suffix}")
            try:
                workbook.save(candidate)
                saved_to = str(candidate)
                break
            except PermissionError:
                continue
        else:
            raise PermissionError(f"Could not save output workbook near {OUT}")

    print("Dataset merged successfully!")
    print("Saved to:", saved_to)


if __name__ == "__main__":
    main()
    main()
