"""
Build the Western Province EV Charging Demand table.

Sources:
  - Population_Table_western.xlsx (sheet A5) -> DS-division population, 2024
  - Western_Province_Data.xlsx               -> EV registrations (raw, by District)
  - Western_Province_EV_Chargers.xlsx        -> existing charger locations (by DS Division)

Output: Western_Province_EV_Demand_Model.xlsx
  - "DS Demand Model" : the merged table the user asked for (live formulas)
  - "EV Registrations" : raw EV registration rows (source for District EV Count)
  - "Chargers Raw"      : raw charger rows (source for Existing Charging Points)

NOTE: Area (km^2) per Divisional Secretariat Division is not present in any of the
three uploaded files, so it cannot be looked up or computed here. The Area column
is left as a highlighted input cell for the user to fill in; every downstream
formula (Density -> Urbanization Proxy -> Demand Score -> City EV Demand ->
Estimated EV Demand -> Recommended Charging Stations) is wired live off that cell,
so the whole table recalculates the moment Area is filled in.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import pandas as pd
import openpyxl as pyxl

SRC_POP = r"C:\Users\ASUS\OneDrive\LMS\3.1\Gruop Project\Population_Tables.xlsx"

SRC_EV = r"C:\Users\ASUS\OneDrive\LMS\3.1\Gruop Project\Western_Province_Data.xlsx"

SRC_CHG = r"C:\Users\ASUS\OneDrive\LMS\3.1\Gruop Project\Western_Province_EV_Chargers.xlsx"

OUT = r"C:\Users\ASUS\OneDrive\LMS\3.1\Gruop Project\Western_Province_EV_Demand_Model.xlsx"

FONT = "Arial"

# ---------------------------------------------------------------------------
# 1. Pull DS-division population from Population_Table_western.xlsx (sheet A5)
# ---------------------------------------------------------------------------
wb_pop = openpyxl.load_workbook(SRC_POP, data_only=True)
ws_pop = wb_pop["A5"]

ds_rows = []  # (District, DS Division, Population)
current_district = None
for r in range(8, 52):
    a = ws_pop.cell(r, 1).value
    b = ws_pop.cell(r, 2).value
    c = ws_pop.cell(r, 3).value
    if a and "District" in str(a):
        current_district = str(a).replace("District", "").strip()
        continue
    if b:
        ds = str(b).replace("\n", "").strip()
        ds_rows.append((current_district, ds, c))

# keep only the three Western Province districts (drop the Sri Lanka total row)
ds_rows = [row for row in ds_rows if row[0] in ("Colombo", "Gampaha", "Kalutara")]
assert len(ds_rows) == 39, f"expected 39 DS divisions, got {len(ds_rows)}"

# ---------------------------------------------------------------------------
# 2. Raw EV registrations (kept as a source sheet -> District EV Count via SUMIF)
# ---------------------------------------------------------------------------
ev_df = pd.read_excel(SRC_EV)
ev_df["District"] = ev_df["District"].astype(str).str.strip().str.title()

# ---------------------------------------------------------------------------
# 3. Raw charger list (kept as a source sheet -> Existing Charging Points via COUNTIF)
# ---------------------------------------------------------------------------
chg_df = pd.read_excel(SRC_CHG, sheet_name="Western Province Chargers", header=1)
chg_df["Divisional Secretary's Division"] = (
    chg_df["Divisional Secretary's Division"].astype(str).str.strip()
)
# fix an obvious source typo so COUNTIF matches the population sheet's DS names
chg_df["Divisional Secretary's Division"] = chg_df[
    "Divisional Secretary's Division"
].replace({"Colomno": "Colombo"})

# ===========================================================================
# BUILD WORKBOOK
# ===========================================================================
wb = openpyxl.Workbook()

# ---------------------------------------------------------------------------
# Sheet: EV Registrations (raw source, used by SUMIF on the main sheet)
# ---------------------------------------------------------------------------
ws_ev = wb.active
ws_ev.title = "EV Registrations"
ev_cols = list(ev_df.columns)
for j, col in enumerate(ev_cols, start=1):
    cell = ws_ev.cell(1, j, col)
    cell.font = Font(name=FONT, bold=True)
for i, row in enumerate(ev_df.itertuples(index=False), start=2):
    for j, val in enumerate(row, start=1):
        ws_ev.cell(i, j, val)
for j, col in enumerate(ev_cols, start=1):
    ws_ev.column_dimensions[get_column_letter(j)].width = max(14, len(col) + 2)
EV_DISTRICT_COL = get_column_letter(ev_cols.index("District") + 1)
EV_COUNT_COL = get_column_letter(ev_cols.index("Count") + 1)
EV_LAST_ROW = ws_ev.max_row

# ---------------------------------------------------------------------------
# Sheet: Chargers Raw (raw source, used by COUNTIF on the main sheet)
# ---------------------------------------------------------------------------
ws_chg = wb.create_sheet("Chargers Raw")
chg_cols = list(chg_df.columns)
for j, col in enumerate(chg_cols, start=1):
    cell = ws_chg.cell(1, j, col)
    cell.font = Font(name=FONT, bold=True)
for i, row in enumerate(chg_df.itertuples(index=False), start=2):
    for j, val in enumerate(row, start=1):
        ws_chg.cell(i, j, val)
for j, col in enumerate(chg_cols, start=1):
    ws_chg.column_dimensions[get_column_letter(j)].width = max(14, len(str(col)) + 2)
CHG_DS_COL = get_column_letter(chg_cols.index("Divisional Secretary's Division") + 1)
CHG_LAST_ROW = ws_chg.max_row

# ---------------------------------------------------------------------------
# Sheet: DS Demand Model (the merged table)
# ---------------------------------------------------------------------------
ws = wb.create_sheet("DS Demand Model", 0)

yellow = PatternFill("solid", fgColor="FFFF00")
header_fill = PatternFill("solid", fgColor="1F4E78")
header_font = Font(name=FONT, bold=True, color="FFFFFF")
title_font = Font(name=FONT, bold=True, size=14)
note_font = Font(name=FONT, italic=True, size=9, color="808080")
spacer_fill = PatternFill("solid", fgColor="D9D9D9")
thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

ws["A1"] = "Western Province EV Charging Demand Model"
ws["A1"].font = title_font

# --- Assumptions block (editable, yellow) ---
ws["A3"] = "Assumptions (edit yellow cells)"
ws["A3"].font = Font(name=FONT, bold=True)

ws["A4"] = "Weight x - Population Share (DS = x*PS + y*US)"
ws["B4"] = 0.5
ws["A5"] = "Weight y - Urbanization Proxy (DS = x*PS + y*US)"
ws["B5"] = 0.5
ws["A6"] = "EVs per Charging Station (capacity assumption)"
ws["B6"] = 50
for addr in ("B4", "B5", "B6"):
    ws[addr].fill = yellow
    ws[addr].font = Font(name=FONT, bold=True)
ws["C6"] = "<- user-editable assumption; not from source data"
ws["C6"].font = note_font
ws["C4"] = "<- must sum to 1.0 for DS to stay on a 0-1 scale"
ws["C4"].font = note_font

HEADER_ROW = 9
FIRST_DATA_ROW = HEADER_ROW + 1
LAST_DATA_ROW = FIRST_DATA_ROW + len(ds_rows) - 1

headers = {
    "A": "District",
    "B": "Divisional Secretariat Division",
    "C": "Population",
    "D": "Area (km2)",
    "E": "Population Share",
    "F": "Density (Pop/km2)",
    "G": "Urbanization Proxy",
    "H": "District EV Count",
    "I": "Demand Score",
    "J": "",
    "K": "City EV Demand",
    "L": "Existing Charging Points (DS)",
    "M": "Estimated EV Demand",
    "N": "",
    "O": "Recommended New Charging Stations",
}
for col, label in headers.items():
    cell = ws[f"{col}{HEADER_ROW}"]
    cell.value = label
    if label:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    else:
        cell.fill = spacer_fill

ws.row_dimensions[HEADER_ROW].height = 45

widths = {
    "A": 12, "B": 24, "C": 12, "D": 12, "E": 15, "F": 16, "G": 16,
    "H": 15, "I": 13, "J": 3, "K": 14, "L": 22, "M": 16, "N": 3, "O": 22,
}
for col, w in widths.items():
    ws.column_dimensions[col].width = w

dv = DataValidation(type="decimal", operator="greaterThan", formula1="0",
                     showErrorMessage=True, errorTitle="Invalid Area",
                     error="Area must be a positive number (km2).")
ws.add_data_validation(dv)

for i, (district, ds, pop) in enumerate(ds_rows):
    r = FIRST_DATA_ROW + i
    ws[f"A{r}"] = district
    ws[f"B{r}"] = ds
    ws[f"C{r}"] = pop
    ws[f"D{r}"] = None  # <- user must fill in Area (km2)
    ws[f"D{r}"].fill = yellow
    dv.add(ws[f"D{r}"])

    ws[f"E{r}"] = f"=C{r}/SUMIF($A${FIRST_DATA_ROW}:$A${LAST_DATA_ROW},A{r},$C${FIRST_DATA_ROW}:$C${LAST_DATA_ROW})"
    ws[f"F{r}"] = f'=IFERROR(C{r}/D{r},"")'
    ws[f"G{r}"] = (
        f'=IFERROR((F{r}-_xlfn.MINIFS($F${FIRST_DATA_ROW}:$F${LAST_DATA_ROW},$A${FIRST_DATA_ROW}:$A${LAST_DATA_ROW},A{r},$F${FIRST_DATA_ROW}:$F${LAST_DATA_ROW},"<>"))'
        f'/(_xlfn.MAXIFS($F${FIRST_DATA_ROW}:$F${LAST_DATA_ROW},$A${FIRST_DATA_ROW}:$A${LAST_DATA_ROW},A{r},$F${FIRST_DATA_ROW}:$F${LAST_DATA_ROW},"<>")'
        f'-_xlfn.MINIFS($F${FIRST_DATA_ROW}:$F${LAST_DATA_ROW},$A${FIRST_DATA_ROW}:$A${LAST_DATA_ROW},A{r},$F${FIRST_DATA_ROW}:$F${LAST_DATA_ROW},"<>")),"")'
    )
    ws[f"H{r}"] = f"=SUMIF('EV Registrations'!${EV_DISTRICT_COL}$2:${EV_DISTRICT_COL}${EV_LAST_ROW},A{r},'EV Registrations'!${EV_COUNT_COL}$2:${EV_COUNT_COL}${EV_LAST_ROW})"
    ws[f"I{r}"] = f'=IFERROR($B$4*E{r}+$B$5*G{r},"")'
    ws[f"K{r}"] = f'=IFERROR(H{r}*I{r}/SUMIF($A${FIRST_DATA_ROW}:$A${LAST_DATA_ROW},A{r},$I${FIRST_DATA_ROW}:$I${LAST_DATA_ROW}),"")'
    ws[f"L{r}"] = f"=COUNTIF('Chargers Raw'!${CHG_DS_COL}$2:${CHG_DS_COL}${CHG_LAST_ROW},B{r})"
    ws[f"M{r}"] = f'=IFERROR(ROUND(K{r},0),"")'
    ws[f"O{r}"] = f'=IFERROR(MAX(0,ROUNDUP(M{r}/$B$6,0)-L{r}),"")'

    for col in "ABCDEFGHIKLMO":
        c = ws[f"{col}{r}"]
        c.font = Font(name=FONT)
        c.border = border
        if col in "CDHL":
            c.number_format = "#,##0"
        elif col in "EGI":
            c.number_format = "0.0%" if col in "E" else "0.000"
        elif col in "KM":
            c.number_format = "#,##0"
        elif col == "O":
            c.number_format = "#,##0"
    ws[f"J{r}"].fill = spacer_fill
    ws[f"N{r}"].fill = spacer_fill

# --- District subtotal rows for readability (formulas, not hardcoded) ---
ws.freeze_panes = f"A{FIRST_DATA_ROW}"

note_row = LAST_DATA_ROW + 2
notes = [
    "Notes / assumptions:",
    "- Population: Dept. of Census & Statistics, 'Population by sex, age and district according to Divisional Secretary's Division, 2024' (Population_Table_western.xlsx, sheet A5).",
    "- Area (km2): NOT present in any supplied file - user input required (yellow cells). All formulas below recalculate automatically once filled in.",
    "- Population Share and the min/max used to normalize Urbanization Proxy are computed within each District (not the whole province), since EV counts and the demand-allocation formula are applied district-by-district.",
    "- Density / Urbanization Proxy / Demand Score follow the slide's formulas: Den_c = Pop_c/Area_c; U_c = (Den_c-min(Den))/(max(Den)-min(Den)); DS = x*PS + y*US.",
    "- District EV Count: sum of 'Count' from Western_Province_Data.xlsx (EV registrations), by District - see 'EV Registrations' sheet.",
    "- City EV Demand: CityEV = DistrictEV x DS/SUM(DS) for DS divisions in that district (province-level formula adapted to district-level, since EV counts are only available by district).",
    "- Existing Charging Points: count of stations per DS Division from Western_Province_EV_Chargers.xlsx - see 'Chargers Raw' sheet ('Colomno' typo corrected to 'Colombo').",
    "- Estimated EV Demand: City EV Demand rounded to a whole number of vehicles.",
    "- Recommended New Charging Stations: MAX(0, ROUNDUP(Estimated EV Demand / EVs-per-station, 0) - Existing Charging Points). 'EVs per Charging Station' is an editable assumption (cell B6), not sourced data.",
]
for i, line in enumerate(notes):
    c = ws.cell(note_row + i, 1, line)
    c.font = note_font if i else Font(name=FONT, bold=True, size=9)

wb.save(OUT)
print("saved", OUT)
print("DS rows:", len(ds_rows))